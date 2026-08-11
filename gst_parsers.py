import csv
import json
import re
from pathlib import Path
from typing import Dict, List, Tuple, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# GST Constants
GSTR_VERSION = "GST3.2.4"
PORTAL_JSON_TYP = "OE"
PORTAL_JSON_HASH = "hash"
PORTAL_B2CS_KEYS = (
    "sply_ty",
    "pos",
    "typ",
    "txval",
    "rt",
    "iamt",
    "camt",
    "samt",
    "csamt",
)

STATE_NAMES = {
    "01": "Jammu & Kashmir",
    "02": "Himachal Pradesh",
    "03": "Punjab",
    "04": "Chandigarh",
    "05": "Uttarakhand",
    "06": "Haryana",
    "07": "Delhi",
    "08": "Rajasthan",
    "09": "Uttar Pradesh",
    "10": "Bihar",
    "11": "Sikkim",
    "12": "Arunachal Pradesh",
    "13": "Nagaland",
    "14": "Manipur",
    "15": "Mizoram",
    "16": "Tripura",
    "17": "Meghalaya",
    "18": "Assam",
    "19": "West Bengal",
    "20": "Jharkhand",
    "21": "Odisha",
    "22": "Chhattisgarh",
    "23": "Madhya Pradesh",
    "24": "Gujarat",
    "25": "Daman & Diu",
    "26": "Dadra & Nagar Haveli & Daman & Diu",
    "27": "Maharashtra",
    "29": "Karnataka",
    "30": "Goa",
    "31": "Lakshadweep",
    "32": "Kerala",
    "33": "Tamil Nadu",
    "34": "Puducherry",
    "35": "Andaman & Nicobar Islands",
    "36": "Telangana",
    "37": "Andhra Pradesh",
    "38": "Ladakh",
}

FLIPKART_STATE_CODES = {
    "IN-AP": "37",
    "IN-AR": "12",
    "IN-AS": "18",
    "IN-BR": "10",
    "IN-CT": "22",
    "IN-CH": "04",
    "IN-DL": "07",
    "IN-GA": "30",
    "IN-GJ": "24",
    "IN-HR": "06",
    "IN-HP": "02",
    "IN-JK": "01",
    "IN-JH": "20",
    "IN-KA": "29",
    "IN-KL": "32",
    "IN-LA": "38",
    "IN-LD": "31",
    "IN-MP": "23",
    "IN-MH": "27",
    "IN-MN": "14",
    "IN-ML": "17",
    "IN-MZ": "15",
    "IN-NL": "13",
    "IN-OR": "21",
    "IN-PY": "34",
    "IN-PB": "03",
    "IN-RJ": "08",
    "IN-SK": "11",
    "IN-TN": "33",
    "IN-TG": "36",
    "IN-TR": "16",
    "IN-UP": "09",
    "IN-UT": "05",
    "IN-WB": "19",
}

# Normalize any state string to code
STATE_NORMALIZED_MAP = {
    "KERALA": "32",
    "TAMILNADU": "33",
    "UTTARPRADESH": "09",
    "CHHATTISGARH": "22",
    "CHATTISGARH": "22",  # Meesho misspelling (single H)
    "ANDHRAPRADESH": "37",
    "RAJASTHAN": "08",
    "UTTARAKHAND": "05",
    "UTTARANCHAL": "05",
    "KARNATAKA": "29",
    "DELHI": "07",
    "WESTBENGAL": "19",
    "GUJARAT": "24",
    "ORISSA": "21",
    "ODISHA": "21",
    "JHARKHAND": "20",
    "BIHAR": "10",
    "TELANGANA": "36",
    "JAMMUKASHMIR": "01",
    "JAMMUANDKASHMIR": "01",
    "HARYANA": "06",
    "MAHARASHTRA": "27",
    "ASSAM": "18",
    "ANDAMANNICOBAR": "35",
    "ANDAMANANDNICOBAR": "35",
    "ANDAMANANDNICOBARISLANDS": "35",
    "ANDAMANNICOBARISLANDS": "35",  # "Andaman & Nicobar Islands" after stripping non-alnum
    "CHANDIGARH": "04",
    "PUNJAB": "03",
    "HIMACHALPRADESH": "02",
    "SIKKIM": "11",
    "ARUNACHALPRADESH": "12",
    "NAGALAND": "13",
    "MANIPUR": "14",
    "MIZORAM": "15",
    "TRIPURA": "16",
    "MEGHALAYA": "17",
    "GOA": "30",
    "LAKSHADWEEP": "31",
    "PUDUCHERRY": "34",
    "PONDICHERRY": "34",
    "LADAKH": "38",
    "DAMANDIU": "25",
    "DADRANAGARHAVELI": "26",
    "DADRADRAANDNAGARHAVELI": "26",
    "DADRANAGARHAVELIDAMANDIU": "26",
    "DADRAANDNAGARHAVELIANDDAMANANDDIU": "26",
}

def rate_to_percent(value: Any) -> float:
    """Amazon Ready-to-File stores GST rate as a fraction (0.05 = 5%).
    Convert to a percentage number (5.0). Values already > 1 are left as-is."""
    try:
        r = float(value or 0.0)
    except (TypeError, ValueError):
        return 0.0
    if 0.0 < r <= 1.0:
        r = r * 100.0
    return round(r, 3)


def split_tax(txval: float, rate: float, pos_code: str, seller_pos: str) -> Tuple[float, float, float]:
    """Derive (IGST, CGST, SGST) from taxable value + rate for a place of supply.
    Intra-state (POS == seller state) splits into CGST/SGST; otherwise all IGST.
    Mirrors how the consolidator recomputes portal tax, keeping summaries accurate."""
    tax = round(txval * rate / 100.0, 2)
    if pos_code == seller_pos:
        camt = round(tax / 2.0, 2)
        samt = round(tax - camt, 2)
        return 0.0, camt, samt
    return tax, 0.0, 0.0


def normalize_state_to_code(state_str: Any) -> Optional[str]:
    if pd.isna(state_str) or not state_str:
        return None
    s = str(state_str).strip().upper()
    # Check if it starts with digits (e.g. "09-Uttar Pradesh")
    match = re.match(r"^(\d{2})", s)
    if match:
        code = match.group(1)
        if code in STATE_NAMES:
            return code
    
    # Check in Flipkart state codes (e.g. "IN-RJ")
    if s in FLIPKART_STATE_CODES:
        return FLIPKART_STATE_CODES[s]
    
    # Strip non-alphanumeric and check in normalized map
    norm_s = "".join(c for c in s if c.isalnum())
    if norm_s in STATE_NORMALIZED_MAP:
        return STATE_NORMALIZED_MAP[norm_s]
        
    return None

class TaxRecord:
    def __init__(
        self,
        platform: str,
        pos_code: str,
        rate: float,
        txval: float,
        iamt: float = 0.0,
        camt: float = 0.0,
        samt: float = 0.0,
        csamt: float = 0.0,
        eco_gstin: str = ""
    ):
        self.platform = platform
        self.pos_code = pos_code
        self.rate = rate
        self.txval = round(txval, 2)
        self.iamt = round(iamt, 2)
        self.camt = round(camt, 2)
        self.samt = round(samt, 2)
        self.csamt = round(csamt, 2)
        self.eco_gstin = eco_gstin.strip().upper() if eco_gstin else ""

class ReportParsers:
    @staticmethod
    def parse_flipkart(excel_path: Path) -> Tuple[str, str, List[TaxRecord]]:
        """Parses Flipkart GSTR return report.xlsx."""
        xl = pd.ExcelFile(excel_path)
        
        # 1. Detect Seller GSTIN
        seller_gstin = ""
        for sheet in ("Section 7(A)(2) in GSTR-1", "Section 7(B)(2) in GSTR-1"):
            if sheet in xl.sheet_names:
                frame = pd.read_excel(xl, sheet_name=sheet)
                if not frame.empty and "GSTIN" in frame.columns:
                    gstin = str(frame["GSTIN"].iloc[0]).strip().upper()
                    if len(gstin) == 15:
                        seller_gstin = gstin
                        break
        if not seller_gstin:
            raise ValueError("Seller GSTIN not found in Flipkart report.")
            
        # 2. Detect Flipkart GSTIN
        flipkart_gstin = "29AAACF1435D1Z5" # Default fallback
        if "Section 3 in GSTR-8" in xl.sheet_names:
            frame = pd.read_excel(xl, sheet_name="Section 3 in GSTR-8")
            if not frame.empty and "GSTIN of Flipkart.Com" in frame.columns:
                gstin = str(frame["GSTIN of Flipkart.Com"].iloc[0]).strip().upper()
                if len(gstin) == 15:
                    flipkart_gstin = gstin
                    
        records: List[TaxRecord] = []
        
        # 3. Read Intrastate B2CS (CGST/SGST)
        sheet_intra = "Section 7(A)(2) in GSTR-1"
        if sheet_intra in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name=sheet_intra)
            for _, row in df.iterrows():
                txval = float(row.get("Aggregate Taxable Value Rs.", 0.0) or 0.0)
                if round(txval, 2) == 0.0:
                    continue
                cgst_rate = float(row.get("CGST %", 0.0) or 0.0)
                sgst_rate = float(row.get("SGST/UT %", 0.0) or 0.0)
                rate = cgst_rate + sgst_rate
                
                cgst_amt = float(row.get("CGST Amount Rs.", 0.0) or 0.0)
                sgst_amt = float(row.get("SGST /UT Amount Rs.", 0.0) or 0.0)
                cess = float(row.get("CESS Amount Rs.", 0.0) or 0.0)
                
                records.append(
                    TaxRecord(
                        platform="Flipkart",
                        pos_code=seller_gstin[:2],
                        rate=rate,
                        txval=txval,
                        iamt=0.0,
                        camt=cgst_amt,
                        samt=sgst_amt,
                        csamt=cess,
                        eco_gstin=flipkart_gstin
                    )
                )
                
        # 4. Read Interstate B2CS (IGST)
        sheet_inter = "Section 7(B)(2) in GSTR-1"
        if sheet_inter in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name=sheet_inter)
            for _, row in df.iterrows():
                txval = float(row.get("Aggregate Taxable Value Rs.", 0.0) or 0.0)
                if round(txval, 2) == 0.0:
                    continue
                state_code = normalize_state_to_code(row.get("Delivered State Code"))
                if not state_code:
                    state_name = row.get("Delivered State (PoS)", "Unknown")
                    raise ValueError(f"Could not resolve Flipkart state POS code for: {state_name}")
                
                rate = float(row.get("IGST %", 0.0) or 0.0)
                igst_amt = float(row.get("IGST Amount Rs.", 0.0) or 0.0)
                cess = float(row.get("CESS Amount Rs.", 0.0) or 0.0)
                
                records.append(
                    TaxRecord(
                        platform="Flipkart",
                        pos_code=state_code,
                        rate=rate,
                        txval=txval,
                        iamt=igst_amt,
                        camt=0.0,
                        samt=0.0,
                        csamt=cess,
                        eco_gstin=flipkart_gstin
                    )
                )
                
        return seller_gstin, flipkart_gstin, records

    @staticmethod
    def parse_amazon(csv_path: Path, amazon_gstin_default: str = "27AAPCA8805L1ZS") -> Tuple[str, str, List[TaxRecord]]:
        """Parses an Amazon report, auto-detecting the format.

        Supports both:
          * MTR (Merchant Tax Report) `.csv` — row-level transactions.
          * GST Ready-to-File `.xlsx` — pre-aggregated GSTR-1 sheets (B2C Small/Large).
        Both produce the same TaxRecord shape so downstream output is identical.
        """
        # xlsx/xls files begin with the ZIP "PK\x03\x04" signature; MTR is plain CSV.
        with open(csv_path, "rb") as fh:
            signature = fh.read(4)
        if signature[:2] == b"PK":
            return ReportParsers.parse_amazon_ready_to_file(csv_path, amazon_gstin_default)
        return ReportParsers._parse_amazon_mtr(csv_path, amazon_gstin_default)

    @staticmethod
    def parse_amazon_ready_to_file(
        excel_path: Path,
        amazon_gstin_default: str = "27AAPCA8805L1ZS",
        sheets: Tuple[str, ...] = ("B2C Small", "B2C Large"),
    ) -> Tuple[str, str, List[TaxRecord]]:
        """Parses Amazon's GST Ready-to-File report (.xlsx).

        By default reads both **B2C Small** and **B2C Large** sheets into B2CS
        records (legacy behaviour). Callers that route B2C Large invoices to
        Table 5 separately (the GST Helper pipeline) MUST pass
        sheets=("B2C Small",) to avoid double-reporting them.
        """
        xl = pd.ExcelFile(excel_path)

        # 1. Seller (Merchant) GSTIN — value sits directly below the label cell.
        seller_gstin = ""
        if "GSTIN" in xl.sheet_names:
            raw = pd.read_excel(xl, sheet_name="GSTIN", header=None)
            for i in range(len(raw)):
                for j in range(raw.shape[1]):
                    cell = str(raw.iat[i, j]).strip()
                    if cell.lower() == "merchant gstin" and i + 1 < len(raw):
                        candidate = str(raw.iat[i + 1, j]).strip().upper()
                        if len(candidate) == 15:
                            seller_gstin = candidate
                        break
                if seller_gstin:
                    break
        if not seller_gstin or len(seller_gstin) != 15:
            raise ValueError(
                "Could not locate the Merchant GSTIN in the Amazon Ready-to-File report "
                "(expected on the 'GSTIN' sheet)."
            )

        seller_pos = seller_gstin[:2]
        records: List[TaxRecord] = []
        representative_eco = amazon_gstin_default

        def read_section(sheet_name: str) -> pd.DataFrame:
            """Locate the real header row (the summary block sits above it)."""
            if sheet_name not in xl.sheet_names:
                return pd.DataFrame()
            probe = pd.read_excel(xl, sheet_name=sheet_name, header=None, nrows=12)
            header_idx = None
            for i in range(len(probe)):
                row_vals = [str(v).strip() for v in probe.iloc[i].tolist()]
                if "Place Of Supply" in row_vals and "Taxable Value" in row_vals:
                    header_idx = i
                    break
            if header_idx is None:
                return pd.DataFrame()
            return pd.read_excel(xl, sheet_name=sheet_name, header=header_idx)

        def eco_from(value: Any) -> str:
            candidate = str(value).strip().upper() if value is not None else ""
            return candidate if len(candidate) == 15 else amazon_gstin_default

        # 2. Selected B2C sheets consolidated into B2CS output.
        for sheet_name in sheets:
            df = read_section(sheet_name)
            if df.empty:
                continue
            for _, row in df.iterrows():
                pos_code = normalize_state_to_code(row.get("Place Of Supply"))
                if not pos_code:
                    continue
                txval = float(row.get("Taxable Value", 0.0) or 0.0)
                if round(txval, 2) == 0.0:
                    continue
                rate = rate_to_percent(row.get("Rate"))
                cess = float(row.get("Cess Amount", 0.0) or 0.0)
                eco_gstin = eco_from(row.get("E-Commerce GSTIN"))
                if representative_eco == amazon_gstin_default and eco_gstin != amazon_gstin_default:
                    representative_eco = eco_gstin
                iamt, camt, samt = split_tax(txval, rate, pos_code, seller_pos)
                records.append(
                    TaxRecord(
                        platform="Amazon",
                        pos_code=pos_code,
                        rate=rate,
                        txval=txval,
                        iamt=iamt,
                        camt=camt,
                        samt=samt,
                        csamt=cess,
                        eco_gstin=eco_gstin,
                    )
                )

        return seller_gstin, representative_eco, records

    @staticmethod
    def _parse_amazon_mtr(csv_path: Path, amazon_gstin_default: str = "27AAPCA8805L1ZS") -> Tuple[str, str, List[TaxRecord]]:
        """Parses Amazon MTR tax report.csv."""
        # Detect encoding
        try:
            df = pd.read_csv(csv_path)
        except Exception:
            df = pd.read_csv(csv_path, encoding="utf-8")
            
        # 1. Detect Seller GSTIN
        seller_gstin = ""
        if "Seller Gstin" in df.columns:
            gstin_series = df["Seller Gstin"].dropna()
            if not gstin_series.empty:
                seller_gstin = str(gstin_series.iloc[0]).strip().upper()
        if not seller_gstin:
            # Try lowercase alternative or check shape
            seller_gstin_cols = [c for c in df.columns if "seller" in c.lower() and "gst" in c.lower()]
            if seller_gstin_cols:
                gstin_series = df[seller_gstin_cols[0]].dropna()
                if not gstin_series.empty:
                    seller_gstin = str(gstin_series.iloc[0]).strip().upper()
        
        if not seller_gstin or len(seller_gstin) != 15:
            raise ValueError(f"Could not locate valid Seller GSTIN in Amazon report. Got: {seller_gstin}")
            
        # Amazon operator GSTIN
        eco_gstin = amazon_gstin_default
        
        records: List[TaxRecord] = []
        
        # Iterate over records (Shipment, Refund, Cancel)
        # We can sum them directly. Refunds and Cancels have correct signed values in columns like Tax Exclusive Gross and Tax Amounts
        for idx, row in df.iterrows():
            t_type = str(row.get("Transaction Type", "")).strip()
            if t_type not in ("Shipment", "Refund"):
                continue # Skip Cancel and other rows (usually zero anyway)
                
            txval = float(row.get("Tax Exclusive Gross", 0.0) or 0.0)
            if round(txval, 2) == 0.0:
                continue
                
            state_val = row.get("Ship To State")
            pos_code = normalize_state_to_code(state_val)
            if not pos_code:
                # Fallback to Bill From State or Skip?
                pos_code = normalize_state_to_code(row.get("Bill From State"))
                if not pos_code:
                    continue # Skip row if state cannot be mapped
            
            # Tax rates
            igst_rate = float(row.get("Igst Rate", 0.0) or 0.0)
            cgst_rate = float(row.get("Cgst Rate", 0.0) or 0.0)
            sgst_rate = float(row.get("Sgst Rate", 0.0) or 0.0)

            # Round: 0.28*100 -> 28.000000000000004 would break POS+rate grouping.
            rate = round((igst_rate + cgst_rate + sgst_rate) * 100.0, 2)
            
            # Taxes (signed fields from Amazon represent refunds as negative)
            igst_amt = float(row.get("Igst Tax", 0.0) or 0.0)
            cgst_amt = float(row.get("Cgst Tax", 0.0) or 0.0)
            sgst_amt = float(row.get("Sgst Tax", 0.0) or 0.0)
            cess_amt = float(row.get("Compensatory Cess Tax", 0.0) or 0.0)
            
            records.append(
                TaxRecord(
                    platform="Amazon",
                    pos_code=pos_code,
                    rate=rate,
                    txval=txval,
                    iamt=igst_amt,
                    camt=cgst_amt,
                    samt=sgst_amt,
                    csamt=cess_amt,
                    eco_gstin=eco_gstin
                )
            )
            
        return seller_gstin, eco_gstin, records

    @staticmethod
    def parse_meesho(sales_excel_path: Path, returns_excel_path: Optional[Path], meesho_gstin_default: str = "09AARCM9332R1CM") -> Tuple[str, str, List[TaxRecord]]:
        """Parses Meesho Sales and Returns reports."""
        xl_sales = pd.ExcelFile(sales_excel_path)
        # Read the first data sheet
        data_sheet = [s for s in xl_sales.sheet_names if s != "Help"][0]
        df_sales = pd.read_excel(xl_sales, sheet_name=data_sheet)
        
        # 1. Detect Seller GSTIN
        seller_gstin = ""
        if "gstin" in df_sales.columns:
            gstin_series = df_sales["gstin"].dropna()
            if not gstin_series.empty:
                seller_gstin = str(gstin_series.iloc[0]).strip().upper()
        if not seller_gstin or len(seller_gstin) != 15:
            raise ValueError("Valid Seller GSTIN not found in Meesho sales report.")
            
        # 2. Detect Meesho GSTIN
        meesho_gstin = meesho_gstin_default
        if "eco_tcs_gstin" in df_sales.columns:
            gstin_series = df_sales["eco_tcs_gstin"].dropna()
            if not gstin_series.empty:
                meesho_gstin = str(gstin_series.iloc[0]).strip().upper()
                
        # Aggregate Sales
        sales_data: Dict[Tuple[str, float], Tuple[float, float]] = {} # (pos, rate) -> (taxable, tax)
        
        for _, row in df_sales.iterrows():
            pos_code = normalize_state_to_code(row.get("end_customer_state_new"))
            if not pos_code:
                # Never skip silently: a dropped row means under-reported tax.
                raise ValueError(
                    f"Could not resolve Meesho state POS code for: {row.get('end_customer_state_new')!r} "
                    "(sales report). Add it to STATE_NORMALIZED_MAP."
                )
            rate = float(row.get("gst_rate", 0.0) or 0.0)
            txval = float(row.get("total_taxable_sale_value", 0.0) or 0.0)
            tax = float(row.get("tax_amount", 0.0) or 0.0)

            key = (pos_code, rate)
            old_txval, old_tax = sales_data.get(key, (0.0, 0.0))
            sales_data[key] = (old_txval + txval, old_tax + tax)
            
        # Aggregate Returns (if provided)
        returns_data: Dict[Tuple[str, float], Tuple[float, float]] = {}
        if returns_excel_path and returns_excel_path.exists():
            xl_ret = pd.ExcelFile(returns_excel_path)
            ret_sheet = [s for s in xl_ret.sheet_names if s != "Help"][0]
            df_ret = pd.read_excel(xl_ret, sheet_name=ret_sheet)
            
            for _, row in df_ret.iterrows():
                pos_code = normalize_state_to_code(row.get("end_customer_state_new"))
                if not pos_code:
                    raise ValueError(
                        f"Could not resolve Meesho state POS code for: {row.get('end_customer_state_new')!r} "
                        "(returns report). Add it to STATE_NORMALIZED_MAP."
                    )
                rate = float(row.get("gst_rate", 0.0) or 0.0)
                txval = float(row.get("total_taxable_sale_value", 0.0) or 0.0)
                tax = float(row.get("tax_amount", 0.0) or 0.0)
                
                key = (pos_code, rate)
                old_txval, old_tax = returns_data.get(key, (0.0, 0.0))
                returns_data[key] = (old_txval + txval, old_tax + tax)
                
        # Calculate Net (Sales - Returns)
        records: List[TaxRecord] = []
        all_keys = set(sales_data.keys()).union(set(returns_data.keys()))
        
        seller_pos = seller_gstin[:2]
        
        for pos_code, rate in all_keys:
            s_val, s_tax = sales_data.get((pos_code, rate), (0.0, 0.0))
            r_val, r_tax = returns_data.get((pos_code, rate), (0.0, 0.0))
            
            net_txval = s_val - r_val
            net_tax = s_tax - r_tax
            
            if round(net_txval, 2) == 0.0:
                continue
                
            # Inter-State vs Intra-State Tax Allocation
            if pos_code == seller_pos:
                # Intra-state: Split tax equally into CGST and SGST
                camt = round(net_tax / 2.0, 2)
                samt = round(net_tax - camt, 2)
                iamt = 0.0
            else:
                # Inter-state: All tax is IGST
                iamt = net_tax
                camt = 0.0
                samt = 0.0
                
            records.append(
                TaxRecord(
                    platform="Meesho",
                    pos_code=pos_code,
                    rate=rate,
                    txval=net_txval,
                    iamt=iamt,
                    camt=camt,
                    samt=samt,
                    csamt=0.0,
                    eco_gstin=meesho_gstin
                )
            )
            
        return seller_gstin, meesho_gstin, records

class GSTConsolidator:
    @staticmethod
    def consolidate(
        flipkart_records: List[TaxRecord],
        amazon_records: List[TaxRecord],
        meesho_records: List[TaxRecord],
        seller_gstin: str
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, Any]]:
        """
        Consolidates reports across platforms.
        Returns:
          1. Consolidated B2CS rows (for Excel display & CSV output)
          2. Portal JSON B2CS rows
          3. Summary breakdown stats
        """
        seller_pos = seller_gstin[:2]
        
        # Combine all records
        all_records = flipkart_records + amazon_records + meesho_records
        
        # Group by platform breakdown
        breakdown_stats = {
            "Flipkart": {"taxable": 0.0, "tax": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "cess": 0.0},
            "Amazon": {"taxable": 0.0, "tax": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "cess": 0.0},
            "Meesho": {"taxable": 0.0, "tax": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "cess": 0.0},
            "Consolidated": {"taxable": 0.0, "tax": 0.0, "cgst": 0.0, "sgst": 0.0, "igst": 0.0, "cess": 0.0}
        }
        
        for r in all_records:
            stats = breakdown_stats[r.platform]
            stats["taxable"] += r.txval
            stats["cgst"] += r.camt
            stats["sgst"] += r.samt
            stats["igst"] += r.iamt
            stats["cess"] += r.csamt
            stats["tax"] += (r.camt + r.samt + r.iamt + r.csamt)
            
            # Consolidated Sums
            cstats = breakdown_stats["Consolidated"]
            cstats["taxable"] += r.txval
            cstats["cgst"] += r.camt
            cstats["sgst"] += r.samt
            cstats["igst"] += r.iamt
            cstats["cess"] += r.csamt
            cstats["tax"] += (r.camt + r.samt + r.iamt + r.csamt)
            
        # Round breakdown stats
        for key in breakdown_stats:
            for field in breakdown_stats[key]:
                breakdown_stats[key][field] = round(breakdown_stats[key][field], 2)
                
        # Group by POS + Rate + E-commerce GSTIN (for offline tool CSV format)
        csv_groups: Dict[Tuple[str, float, str], Dict[str, float]] = {}
        for r in all_records:
            key = (r.pos_code, r.rate, r.eco_gstin)
            group = csv_groups.get(key, {"txval": 0.0, "cess": 0.0})
            group["txval"] += r.txval
            group["cess"] += r.csamt
            csv_groups[key] = group
            
        csv_rows: List[Dict[str, Any]] = []
        for (pos_code, rate, eco_gstin), vals in csv_groups.items():
            txval = vals["txval"]
            if round(txval, 2) == 0.0:
                continue
            pos_label = f"{pos_code}-{STATE_NAMES[pos_code]}"
            rate_label = f"{rate:.1f}".rstrip("0").rstrip(".")
            csv_rows.append({
                "Type": "E",
                "Place Of Supply": pos_label,
                "Rate": rate_label,
                "Applicable % of Tax Rate": "",
                "Taxable Value": round(txval, 2),
                "Cess Amount": round(vals["cess"], 2),
                "E-Commerce GSTIN": eco_gstin,
                "pos_code": pos_code,
                "rate": rate
            })
            
        csv_rows.sort(key=lambda x: (x["pos_code"], x["rate"]))
        
        # Group by POS + Rate + Supply Type (for portal JSON format, no E-commerce GSTIN)
        json_groups: Dict[Tuple[str, float, str], Dict[str, float]] = {}
        for r in all_records:
            sply_ty = "INTRA" if r.pos_code == seller_pos else "INTER"
            key = (r.pos_code, r.rate, sply_ty)
            group = json_groups.get(key, {
                "txval": 0.0,
                "iamt": 0.0,
                "camt": 0.0,
                "samt": 0.0,
                "csamt": 0.0
            })
            group["txval"] += r.txval
            group["iamt"] += r.iamt
            group["camt"] += r.camt
            group["samt"] += r.samt
            group["csamt"] += r.csamt
            json_groups[key] = group
            
        json_rows: List[Dict[str, Any]] = []
        for (pos_code, rate, sply_ty), ams in json_groups.items():
            txval = ams["txval"]
            if round(txval, 2) == 0.0:
                continue
                
            # Direct math checks
            total_tx = round(txval * rate / 100.0, 2)
            if sply_ty == "INTRA":
                camt = round(total_tx / 2.0, 2)
                samt = round(total_tx - camt, 2)
                iamt = 0.0
            else:
                iamt = total_tx
                camt = 0.0
                samt = 0.0
                
            # Convert numbers to portal-friendly (integer if whole, otherwise float)
            def num(v: float) -> Any:
                rv = round(v, 2)
                return int(rv) if rv == int(rv) else rv
                
            json_rows.append({
                "sply_ty": sply_ty,
                "pos": pos_code,
                "typ": PORTAL_JSON_TYP,
                "txval": num(txval),
                "rt": num(rate),
                "iamt": num(iamt),
                "camt": num(camt),
                "samt": num(samt),
                "csamt": num(ams["csamt"]),
            })
            
        json_rows.sort(key=lambda x: (x["sply_ty"], x["pos"], x["rt"]))
        
        return csv_rows, json_rows, breakdown_stats

    @staticmethod
    def generate_excel(
        seller_gstin: str,
        period: str,
        csv_rows: List[Dict[str, Any]],
        stats: Dict[str, Any],
        flipkart_records: List[TaxRecord],
        amazon_records: List[TaxRecord],
        meesho_records: List[TaxRecord],
        excel_path: Path
    ) -> None:
        """Generates a beautifully styled, professional, multi-sheet Excel workbook."""
        wb = Workbook()
        
        # Color Palette - Modern Emerald/Navy
        header_fill = PatternFill(start_color="1B365D", end_color="1B365D", fill_type="solid") # Dark Navy
        accent_fill = PatternFill(start_color="E6F0FA", end_color="E6F0FA", fill_type="solid") # Very Light Blue
        zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid") # Light grayish blue
        total_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Slate light
        
        # Fonts
        title_font = Font(name="Segoe UI", size=16, bold=True, color="1B365D")
        section_font = Font(name="Segoe UI", size=12, bold=True, color="000000")
        header_font = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
        bold_font = Font(name="Segoe UI", size=10, bold=True, color="000000")
        regular_font = Font(name="Segoe UI", size=10, color="333333")
        
        # Borders
        thin_line = Side(border_style="thin", color="CBD5E1")
        double_line = Side(border_style="double", color="475569")
        grid_border = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)
        total_border = Border(top=thin_line, bottom=double_line)
        
        # Alignments
        left_align = Alignment(horizontal="left", vertical="center")
        center_align = Alignment(horizontal="center", vertical="center")
        right_align = Alignment(horizontal="right", vertical="center")
        
        # Number formats
        currency_format = '₹#,##0.00'
        percentage_format = '0.0%'
        
        # ----------------------------------------------------
        # SHEET 1: DASHBOARD
        # ----------------------------------------------------
        ws1 = wb.active
        ws1.title = "Dashboard"
        ws1.views.sheetView[0].showGridLines = True
        
        # Title block
        ws1["A1"] = "GST FILING E-COMMERCE CONSOLIDATION REPORT"
        ws1["A1"].font = title_font
        ws1["A2"] = f"Seller GSTIN: {seller_gstin}   |   Filing Period: {period}"
        ws1["A2"].font = Font(name="Segoe UI", size=11, italic=True, color="475569")
        
        # Platforms Summary Table
        ws1["A4"] = "Platform-wise Performance Summary"
        ws1["A4"].font = section_font
        
        summary_headers = ["Platform", "Net Taxable Value (₹)", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)", "Total Tax (₹)"]
        for col_idx, h in enumerate(summary_headers, start=1):
            cell = ws1.cell(row=5, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = grid_border
            
        platforms = ["Flipkart", "Amazon", "Meesho", "Consolidated"]
        for row_idx, plat in enumerate(platforms, start=6):
            plat_stats = stats[plat]
            is_total = (plat == "Consolidated")
            
            row_data = [
                f"{plat} Total" if is_total else plat,
                plat_stats["taxable"],
                plat_stats["igst"],
                plat_stats["cgst"],
                plat_stats["sgst"],
                plat_stats["cess"],
                plat_stats["tax"]
            ]
            
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=val)
                cell.font = bold_font if is_total else regular_font
                cell.alignment = left_align if col_idx == 1 else right_align
                cell.border = total_border if is_total else grid_border
                
                if is_total:
                    cell.fill = total_fill
                elif row_idx % 2 == 1:
                    cell.fill = zebra_fill
                    
                if col_idx > 1:
                    cell.number_format = currency_format
                    
        # ----------------------------------------------------
        # SHEET 2: CONSOLIDATED B2CS (OFFLINE TOOL CSV)
        # ----------------------------------------------------
        ws2 = wb.create_sheet(title="Consolidated B2CS")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2["A1"] = "GSTR-1 B2CS Consolidated Upload Sheet"
        ws2["A1"].font = title_font
        ws2["A2"] = "Formatted precisely according to the GST Returns Offline Tool specifications."
        ws2["A2"].font = Font(name="Segoe UI", size=10, italic=True, color="475569")
        
        headers = ["Type", "Place Of Supply", "Rate", "Applicable % of Tax Rate", "Taxable Value", "Cess Amount", "E-Commerce GSTIN"]
        for col_idx, h in enumerate(headers, start=1):
            cell = ws2.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = grid_border
            
        total_taxable = 0.0
        total_cess = 0.0
        
        for row_idx, r in enumerate(csv_rows, start=5):
            total_taxable += r["Taxable Value"]
            total_cess += r["Cess Amount"]
            
            row_data = [
                r["Type"],
                r["Place Of Supply"],
                float(r["Rate"]) / 100.0 if isinstance(r["Rate"], (int, float)) or (isinstance(r["Rate"], str) and r["Rate"].replace(".","").isdigit()) else r["Rate"],
                r["Applicable % of Tax Rate"],
                r["Taxable Value"],
                r["Cess Amount"],
                r["E-Commerce GSTIN"]
            ]
            
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=val)
                cell.font = regular_font
                cell.border = grid_border
                
                if row_idx % 2 == 1:
                    cell.fill = zebra_fill
                    
                if col_idx in (1, 2, 4, 7):
                    cell.alignment = center_align if col_idx != 2 else left_align
                elif col_idx == 3:
                    cell.alignment = right_align
                    try:
                        cell.value = float(val)
                        if cell.value > 1:
                            cell.value = cell.value / 100.0
                        cell.number_format = percentage_format
                    except Exception:
                        pass
                else:
                    cell.alignment = right_align
                    cell.number_format = currency_format
                    
        # Total Row for Consolidated B2CS
        tot_row = len(csv_rows) + 5
        ws2.cell(row=tot_row, column=1, value="Total").font = bold_font
        ws2.cell(row=tot_row, column=1).alignment = left_align
        ws2.cell(row=tot_row, column=1).border = total_border
        ws2.cell(row=tot_row, column=1).fill = total_fill
        
        for col in range(2, 8):
            cell = ws2.cell(row=tot_row, column=col)
            cell.border = total_border
            cell.fill = total_fill
            if col == 5:
                cell.value = total_taxable
                cell.font = bold_font
                cell.alignment = right_align
                cell.number_format = currency_format
            elif col == 6:
                cell.value = total_cess
                cell.font = bold_font
                cell.alignment = right_align
                cell.number_format = currency_format
                
        # ----------------------------------------------------
        # SHEET 3: PLATFORM BREAKDOWNS
        # ----------------------------------------------------
        ws3 = wb.create_sheet(title="Platform Breakdowns")
        ws3.views.sheetView[0].showGridLines = True
        
        ws3["A1"] = "Detailed Platform Breakdowns (Net Aggregates)"
        ws3["A1"].font = title_font
        
        plat_headers = ["Platform", "Place of Supply", "Rate (%)", "Taxable Value (₹)", "IGST (₹)", "CGST (₹)", "SGST (₹)", "Cess (₹)", "E-Commerce GSTIN"]
        for col_idx, h in enumerate(plat_headers, start=1):
            cell = ws3.cell(row=4, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = grid_border
            
        all_breakdowns = [
            ("Flipkart", flipkart_records),
            ("Amazon", amazon_records),
            ("Meesho", meesho_records)
        ]
        
        current_row = 5
        for plat_name, records in all_breakdowns:
            if not records:
                continue
            for r in records:
                row_data = [
                    plat_name,
                    f"{r.pos_code}-{STATE_NAMES.get(r.pos_code, 'Unknown')}",
                    r.rate / 100.0,
                    r.txval,
                    r.iamt,
                    r.camt,
                    r.samt,
                    r.csamt,
                    r.eco_gstin
                ]
                for col_idx, val in enumerate(row_data, start=1):
                    cell = ws3.cell(row=current_row, column=col_idx, value=val)
                    cell.font = regular_font
                    cell.border = grid_border
                    
                    if current_row % 2 == 1:
                        cell.fill = zebra_fill
                        
                    if col_idx in (1, 2, 9):
                        cell.alignment = left_align if col_idx != 9 else center_align
                    elif col_idx == 3:
                        cell.alignment = right_align
                        cell.number_format = percentage_format
                    else:
                        cell.alignment = right_align
                        cell.number_format = currency_format
                current_row += 1
                
        # Set column widths nicely for all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                for cell in col:
                    val_str = str(cell.value or '')
                    if cell.number_format == percentage_format and isinstance(cell.value, (int, float)):
                        val_str = f"{cell.value * 100:.1f}%"
                    elif cell.number_format == currency_format and isinstance(cell.value, (int, float)):
                        val_str = f"₹{cell.value:,.2f}"
                    max_len = max(max_len, len(val_str))
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        wb.save(excel_path)
